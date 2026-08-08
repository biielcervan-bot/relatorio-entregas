import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="Gestão de Entregas", layout="wide")
st.title("🚚 Painel Operacional - Relatório de Entregas (Padronizado)")
st.markdown("---")

# Upload de múltiplos arquivos
uploaded_files = st.file_uploader(
    "Faça o upload das planilhas de entregas (.csv ou .xlsx)", 
    type=["csv", "xlsx"], 
    accept_multiple_files=True
)

if uploaded_files:
    try:
        lista_dfs = []
        
        for file in uploaded_files:
            df_temp = None
            if file.name.lower().endswith('.csv'):
                for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1', 'utf-8-sig']:
                    for sep in [';', ',', '\t']:
                        try:
                            file.seek(0)
                            temp = pd.read_csv(file, sep=sep, encoding=enc, low_memory=False, on_bad_lines='skip')
                            if temp is not None and len(temp.columns) > 1:
                                df_temp = temp
                                break
                        except Exception:
                            continue
                    if df_temp is not None:
                        break
            else:
                df_temp = pd.read_excel(file)

            if df_temp is not None:
                lista_dfs.append(df_temp)

        if lista_dfs:
            with st.spinner("🚀 Processando e padronizando dados de entregas na estrutura de leitura..."):
                df = pd.concat(lista_dfs, ignore_index=True)
                
                # Normalização de nomes de colunas
                df.columns = df.columns.astype(str).str.strip().str.upper()

                def buscar_coluna(opcoes):
                    for op in opcoes:
                        if op in df.columns:
                            return op
                    return None

                # Mapeamento Flexível das colunas existentes na base
                col_dt_real = buscar_coluna(['DATA_HORA_APROXIMADA', 'DATA_ENTREGA', 'DATA_HORA', 'DT_ENTREGA', 'DATA_REAL', 'DT_INI_ACAO'])
                col_base = buscar_coluna(['NOM_BASE_OPERACIONAL', 'BASE_OPERACIONAL', 'BASE'])
                col_mun = buscar_coluna(['NOM_MUNICIPIO', 'MUNICIPIO', 'CIDADE'])
                col_unidade = buscar_coluna(['NOM_UNIDADE_LEITURA', 'UNIDADE_LEITURA', 'UNIDADE'])
                col_cod_agente = buscar_coluna(['COD_AGENTE_COMERCIAL', 'COD_AGENTE', 'CODIGO_AGENTE', 'COD_LEITOR'])
                col_zona = buscar_coluna(['ZONA', 'AREA', 'TIPO_AREA', 'URBANA_RURAL', 'LOCALIZACAO', 'SITUACAO'])
                col_tarefa = buscar_coluna(['SEQ_TAREFA', 'COD_TAREFA', 'ID_ENTREGA', 'NOTIFICACAO', 'OS'])

                # 1. Tratamento de Data e Hora
                if col_dt_real:
                    df['DATA_HORA_DT'] = pd.to_datetime(df[col_dt_real], dayfirst=True, errors='coerce')
                    df['DATA_REAL'] = df['DATA_HORA_DT'].dt.strftime('%d/%m/%Y').fillna('Sem Data')
                    df['HORA'] = df['DATA_HORA_DT'].dt.strftime('%H:%M').fillna('N/A')
                else:
                    df['DATA_HORA_DT'] = pd.NaT
                    df['DATA_REAL'] = 'Sem Data'
                    df['HORA'] = 'N/A'

                # 2. Regras Específicas Solicitadas:
                df['BASE_STD'] = df[col_base].fillna('N/A').astype(str).str.strip() if col_base else 'N/A'
                df['MUNICIPIO_STD'] = df[col_mun].fillna('N/A').astype(str).str.strip() if col_mun else 'N/A'
                df['UNIDADE_STD'] = df[col_unidade].fillna('N/A').astype(str).str.strip() if col_unidade else 'N/A'
                df['COD_AGENTE_STD'] = df[col_cod_agente].fillna('N/A').astype(str).str.strip().str.replace(r'\.0$', '', regex=True) if col_cod_agente else 'N/A'
                df['ZONA_STD'] = df[col_zona].fillna('N/A').astype(str).str.strip() if col_zona else 'N/A'
                df['TAREFA_STD'] = df[col_tarefa] if col_tarefa else df.index

                # REGRAS FIXAS DAS INSTRUÇÕES:
                df['LOTE_STD'] = ""            # Lote vem vazio
                df['TIPO_STD'] = "E"           # Tipo fixado como 'E'
                df['NOM_AGENTE_STD'] = ""      # Nome do agente vem vazio
                df['IMP_GRUPO_1'] = 0          # Impedimentos G1 = 0
                df['IMP_GRUPO_2'] = 0          # Impedimentos G2 = 0
                df['QTD_FOTO_NUM'] = 0         # Fotos = 0
                df['LEITURA_LIMPA'] = 1        # Cada entrega é 100% limpa (sem impedimentos)

                df['AGENTE_COMPLETO'] = df['COD_AGENTE_STD']

            # 3. Barra Lateral - Filtros
            st.sidebar.header("🎯 Filtros")

            def criar_multiselect(label, col_name):
                opcoes = sorted([x for x in df[col_name].unique() if str(x) not in ['nan', 'N/A', 'Sem Data', '']])
                if not opcoes:
                    opcoes = sorted([x for x in df[col_name].unique() if str(x) != 'nan'])
                return st.sidebar.multiselect(label, options=opcoes, default=opcoes)

            f_base = criar_multiselect("Base Operacional", 'BASE_STD')
            f_mun = criar_multiselect("Município", 'MUNICIPIO_STD')
            f_unidade = criar_multiselect("Unidade de Leitura", 'UNIDADE_STD')
            f_zona = criar_multiselect("Zona / Área", 'ZONA_STD')
            f_agente = criar_multiselect("Código do Agente Comercial", 'AGENTE_COMPLETO')
            f_data_real = criar_multiselect("Data da Entrega", 'DATA_REAL')

            # Aplicação dos Filtros
            df_filtrado = df[
                (df['BASE_STD'].isin(f_base)) &
                (df['MUNICIPIO_STD'].isin(f_mun)) &
                (df['UNIDADE_STD'].isin(f_unidade)) &
                (df['ZONA_STD'].isin(f_zona)) &
                (df['AGENTE_COMPLETO'].isin(f_agente)) &
                (df['DATA_REAL'].isin(f_data_real))
            ]

            if df_filtrado.empty:
                st.warning("⚠️ Nenhum registro encontrado para os filtros selecionados.")
            else:
                # 4. Indicadores Principais (Estrutura idêntica ao de leitura)
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Total de Entregas", f"{len(df_filtrado):,}".replace(",", "."))
                col2.metric("Entregas Limpas", f"{len(df_filtrado):,}".replace(",", "."))
                col3.metric("Impedimentos G1", "0")
                col4.metric("Impedimentos G2", "0")
                col5.metric("Total de Fotos", "0")

                st.markdown("---")

                # Funções de hora
                def hora_min_max(s, tipo):
                    v = s.dropna()
                    if v.empty: return "N/A"
                    res = v.min() if tipo == 'min' else v.max()
                    return res.strftime('%H:%M') if pd.notna(res) else "N/A"

                # 5. Agrupamento Consolidado Estruturado Igual ao de Leitura
                df_resumo = df_filtrado.groupby([
                    'DATA_REAL', 'BASE_STD', 'MUNICIPIO_STD', 'LOTE_STD', 'UNIDADE_STD',
                    'ZONA_STD', 'TIPO_STD', 'COD_AGENTE_STD', 'NOM_AGENTE_STD'
                ], as_index=False).agg(
                    TOTAL_ENTREGAS=('TAREFA_STD', 'count'),
                    ENTREGAS_LIMPAS=('LEITURA_LIMPA', 'sum'),
                    IMP_G1=('IMP_GRUPO_1', 'sum'),
                    IMP_G2=('IMP_GRUPO_2', 'sum'),
                    TOTAL_FOTOS=('QTD_FOTO_NUM', 'sum'),
                    HORA_INI=('DATA_HORA_DT', lambda x: hora_min_max(x, 'min')),
                    HORA_FIM=('DATA_HORA_DT', lambda x: hora_min_max(x, 'max'))
                )

                df_resumo.columns = [
                    'Data Realização', 'Base Operacional', 'Município', 'Lote',
                    'Unidade de Leitura', 'Zona / Área', 'Tipo Serviço',
                    'Código Agente', 'Nome Agente',
                    'Total Leituras (Entregas)', 'Leituras Limpas (Entregas)', 
                    'Impedimentos G1', 'Impedimentos G2',
                    'Total Fotos', '1ª Entrega', 'Última Entrega'
                ]

                # 6. Gráficos Visualizadores
                col_graf1, col_graf2 = st.columns(2)
                with col_graf1:
                    st.subheader("🏙️ Entregas por Cidade")
                    df_cidade = df_filtrado.groupby('MUNICIPIO_STD').size().reset_index(name='Qtd Entregas')
                    fig_cidade = px.bar(df_cidade, x='MUNICIPIO_STD', y='Qtd Entregas', text_auto=True, 
                                        labels={'MUNICIPIO_STD': 'Município'}, color_discrete_sequence=['#1f77b4'])
                    st.plotly_chart(fig_cidade, use_container_width=True)

                with col_graf2:
                    st.subheader("🏢 Entregas por Base Operacional")
                    df_base_graf = df_filtrado.groupby('BASE_STD').size().reset_index(name='Qtd Entregas')
                    fig_base = px.bar(df_base_graf, x='BASE_STD', y='Qtd Entregas', text_auto=True, 
                                       labels={'BASE_STD': 'Base Operacional'}, color_discrete_sequence=['#2ca02c'])
                    st.plotly_chart(fig_base, use_container_width=True)

                st.markdown("---")
                st.subheader("📋 Resumo Consolidado de Entregas (Estrutura Unificada)")
                st.dataframe(df_resumo, use_container_width=True)

                # 7. GERADOR EXCEL FORMATADO (2 ABAS)
                buffer = io.BytesIO()
                wb = openpyxl.Workbook()
                
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                thin_border = Side(border_style="thin", color="D3D3D3")
                border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                align_center = Alignment(horizontal="center", vertical="center")
                align_right = Alignment(horizontal="right", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")

                # Aba 1: Resumo Consolidado
                ws1 = wb.active
                ws1.title = "Resumo Consolidado"
                
                ws1.append(list(df_resumo.columns))
                for row in df_resumo.itertuples(index=False):
                    ws1.append(list(row))

                for cell in ws1[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center

                for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(df_resumo.columns)):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_left

                for col in ws1.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

                # Aba 2: Base Filtrada Detalhada
                ws2 = wb.create_sheet(title="Base Filtrada Detalhada")
                df_detalhado_export = df_filtrado[[
                    'DATA_REAL', 'BASE_STD', 'MUNICIPIO_STD', 'LOTE_STD', 'UNIDADE_STD',
                    'ZONA_STD', 'TIPO_STD', 'COD_AGENTE_STD', 'NOM_AGENTE_STD', 'HORA',
                    'LEITURA_LIMPA', 'IMP_GRUPO_1', 'IMP_GRUPO_2', 'QTD_FOTO_NUM'
                ]].copy()

                df_detalhado_export.columns = [
                    'Data Realização', 'Base Operacional', 'Município', 'Lote', 'Unidade de Leitura',
                    'Zona / Área', 'Tipo Serviço', 'Código Agente', 'Nome Agente', 'Hora Entrega',
                    'Entrega Limpa (1/0)', 'Impedimento G1', 'Impedimento G2', 'Qtd Fotos'
                ]

                ws2.append(list(df_detalhado_export.columns))
                for row in df_detalhado_export.itertuples(index=False):
                    ws2.append(list(row))

                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center

                for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(df_detalhado_export.columns)):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_left

                for col in ws2.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

                wb.save(buffer)

                st.download_button(
                    label="📥 Baixar Planilha Consolidada de Entregas (Excel Padronizado)",
                    data=buffer.getvalue(),
                    file_name="resumo_entregas_padronizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Erro ao processar os arquivos: {e}")
else:
    st.info("👆 Faça o upload das planilhas de entregas para iniciar.")
